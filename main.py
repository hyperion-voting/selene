import multiprocessing
from pathlib import Path
import argparse
import time

from openpyxl import load_workbook, Workbook
from gmpy2 import powmod, invert, f_mod, mul, mpz, divm
from texttable import Texttable
import threshold_crypto as tc


from group import DHGroup, pars_2048
from primitives import ElGamalEncryption
from parties import Voter, Teller, ElectionAuthority
from util import (
    multi_dim_index,
    print_bb,
    find_entry_by_comm,
    calculate_voter_term,
)

# Note: The variable names used here may not reflect the actual names
# used in the paper since we have renamed them for PEP-8 compliance.
# However, dictionary keys and messages use names from the paper
# for brevity.


parser = argparse.ArgumentParser(
    description="Selene"
)
parser.add_argument(
    "voter_count", metavar="N", type=int, help="Number of voters"
)

parser.add_argument(
    "teller_count", metavar="T", type=int, help="Number of tellers"
)

parser.add_argument(
    "teller_threshold_count",
    metavar="K",
    type=int,
    help="Teller threshold value",
)

parser.add_argument(
    "-maxv",
    "--max-vote",
    metavar="MAX",
    type=str,
    help="Maximum vote value [Default: 1]",
)

parser.add_argument(
    "-v",
    "--verbose",
    action="store_true",
    default=False,
    help="Display bulletin board contents and commitment values (experimental)",
)


args = parser.parse_args()

num_voters = 50
if (
    args.voter_count is not None
    and int(args.voter_count) > 0
    and int(args.voter_count) < 10000
):
    num_voters = int(args.voter_count)
num_tellers = 5
if (
    args.teller_count is not None
    and int(args.teller_count) > 0
    and int(args.teller_count) < 100
):
    num_tellers = int(args.teller_count)
k = 3
if (
    args.teller_threshold_count is not None
    and int(args.teller_threshold_count) > 0
    and int(args.teller_threshold_count) < 100
):
    if int(args.teller_threshold_count) > num_tellers:
        print(
            "The teller threshold value must be less than the number of tellers."
        )
        exit()

vote_min = 0
vote_max = 2
if args.max_vote is not None and int(args.max_vote) > 1:
    vote_max = int(args.max_vote)

q1 = multiprocessing.Queue()
q2 = multiprocessing.Queue()

t_voting_single = 0
t_verification_single = 0
t_re_enc_mix_ver = 0
t_mixing = 0
t_decryption = 0

voters = []
tellers = []

bb = []
final_bb = []
verification_bb = []

teller_sk = []
teller_public_key = ""

teller_registry = []

key_params = pars_2048()
group = DHGroup(key_params.p, key_params.g, key_params.q)

t_voting_single = 0
t_verification_single = 0


def print_bb():
    """Prints out the contents of a bulletin board.
    """
    for item in bb:
        print(item)


def find_entry_by_id(id, items):
    """Returns an entry from a given list of entries

    Args:
        id (str): The value of the 'id' field of the entry to return

    Returns:
        The entry if found, None otherwise.
    """
    for item in items:
        if item.id == id:
            return item
    return None


def find_index_by_id(id, items):
    """Returns an entry from a given list of entries

    Args:
        id (str): The value of the 'id' field of the entry to return

    Returns:
        The entry if found, None otherwise.
    """
    i = 0
    for item in items:
        if item["id"] == id:
            return i
        i = i + 1
    return None


def find_entry_by_tracker(tracker, items):
    """Returns an entry from a given list of entries

    Args:
        comm (str): The value of the 'comm' field of the entry to return

    Returns:
        The entry if found, None otherwise.
    """
    for item in items:
        if item["tracker"] == tracker:
            return item
    return None


vote_map = []


def poc_setup():
    """Sets up voter IDs and voter objects for 'vote_max' voters.
    Generates DSA key pairs for each voter.
    Picks a random vote value for each voter in the range
    ('vote_min':'vote_max').
    Adds all 'voter' objects to the 'voters' list.
    """
    for i in range(vote_min, vote_max + 1):
        g_vote = group.raise_g(i)
        vote_map.append({"vote": i, "g_vote": g_vote})

    for i in range(0, num_voters):
        id = "VT" + str(i)
        voter = Voter(group, id, vote_min, vote_max)
        voter.generate_dsa_keys()
        voter.choose_vote_value()
        voter.generate_trapdoor_keypair()
        # voter.generate_pok_trapdoor_keypair()
        voters.append(voter)


def setup():
    """The setup phase of the protocol.
    Sets up 'num_tellers' teller objects.
    The teller public key and the threshold secret keys for 'num_tellers'
    tally tellers are established.
    Adds all 'teller' objects to the 'tellers' list.
    """
    global teller_public_key
    global teller_sk
    global g_trackers
    global final_bb
    teller_public_key, teller_sk = Teller.generate_threshold_keys(
        k, num_tellers, pars_2048()
    )
    for i in range(0, num_tellers):
        teller = Teller(group, teller_sk[i], teller_public_key)
        tellers.append(teller)

    ea = ElectionAuthority(group)
    trackers = ea.generate_trackers(8, num_voters)
    g_trackers = ea.raise_trackers(trackers)
    encrypted_trackers = ea.encrypt_trackers(teller_public_key.g_a, g_trackers)
    for teller in tellers:
        encrypted_trackers = teller.rencryption_mix_trackers(
            encrypted_trackers
        )

    i = 0
    for voter in voters:
        tracker_voter_pairs.append(
            [voter.id, voter.public_trapdoor_key, encrypted_trackers[i]]
        )
        i = i + 1

    teller_commitments = []
    for teller in tellers:
        teller_commitments.append(
            teller.generate_tracker_commitments(tracker_voter_pairs)
        )

    final_bb = []

    for voter in range(len(teller_commitments[0])):
        temp_h_ri = [mpz(1), mpz(1)]
        temp_v_id = teller_commitments[0][voter]["id"]
        for column in range(len(teller_commitments)):
            temp_h_ri[0] = f_mod(
                mul(
                    temp_h_ri[0],
                    teller_commitments[column][voter]["enc_h_ri"][0],
                ),
                group.p,
            )
            temp_h_ri[1] = f_mod(
                mul(
                    temp_h_ri[1],
                    teller_commitments[column][voter]["enc_h_ri"][1],
                ),
                group.p,
            )
            # abort if v_id doesnt match

        temp_h_ri_enc_tracker = [
            f_mod(
                mul(temp_h_ri[0], tracker_voter_pairs[voter][2][0]), group.p
            ),
            f_mod(
                mul(temp_h_ri[1], tracker_voter_pairs[voter][2][1]), group.p
            ),
        ]
        # print(temp_h_ri_enc_tracker)
        final_bb.append(
            {
                "id": temp_v_id,
                "public_key": "",
                "enc_tracker": tracker_voter_pairs[voter][2],
                "comm": temp_h_ri_enc_tracker,
                "vote": "",
            }
        )
    reconstruct_shares = []
    for teller in tellers:
        reconstruct_shares.append(teller.secret_key_share)
    ege = ElGamalEncryption(group)
    for i in range(0, len(final_bb)):
        partial_decryptions_comm = [
            ege.partial_decrypt(final_bb[i]["comm"], share)
            for share in reconstruct_shares
        ]
        ciphertext_tc_comm = tc.EncryptedMessage(
            final_bb[i]["comm"][0], final_bb[i]["comm"][1], ""
        )
        decrypted_message_comm = ege.threshold_decrypt(
            partial_decryptions_comm,
            ciphertext_tc_comm,
            tc.ThresholdParameters(k, num_tellers),
            pars_2048(),
        )
        final_bb[i]["comm"] = decrypted_message_comm


def voting():
    """The voting phase of the protocol.
    For each 'voter' in the 'voters' list:
        the vote is encrypted under the tellers' threshold public key,
        a proof of wellformedness of the ballot is generated,
        the signed, encrypted ballot is posted to a bulletin board.
    """
    global final_bb
    for voter in voters:
        index = find_index_by_id(voter.id, final_bb)
        t_voting_single_start = time.time()
        voter.encrypt_vote(teller_public_key)
        voter.generate_wellformedness_proof(teller_public_key)
        bb_data = voter.sign_ballot()
        final_bb[index]["vote"] = bb_data
        t_voting = time.time() - t_voting_single_start
        global t_voting_single
        t_voting_single = t_voting_single + t_voting


def tallying():
    """The tallying phase of the protocol.
    The encrypted votes and 'h_r' tuples are subjected to a series of
    parallel rencryption mixes by the tally tellers.
    The tuples are decrypted by the tally tellers and posted to
    a final bulletin board. The code in this phase has been modified to 
    allow it to run faster on a multi-core system.
    """
    global final_bb
    global bb
    global t_mixing
    global t_decryption
    t_mixing_start = time.time()

    tagged_bb = []
    combined_bb = []
    raised_bb = []
    index = 0
    global verification_bb
    temp = []
    for item in final_bb:
        temp_2 = []
        temp_2.append(item["vote"]["ev"])
        temp_2.append(item["enc_tracker"])
        temp.append(temp_2)
        

    previous = temp
    global t_re_enc_mix_ver
    for teller in tellers:
        proof = teller.re_encryption_mix(previous)
        new_list = proof[0]
        t_re_enc_mix_ver_start = time.time()
        teller.verify_re_enc_mix(previous, proof)
        t_re_enc_mix_ver_end = time.time()
        t_re_enc_mix_ver = t_re_enc_mix_ver + (
            t_re_enc_mix_ver_end - t_re_enc_mix_ver_start
        )
        previous = new_list
    t_mixing = (time.time() - t_mixing_start) - t_re_enc_mix_ver
    t_decryption_start = time.time()
    tagged_ciphertexts = teller.tag_ciphertexts(new_list)
    split_ciphertexts = teller.ciphertext_list_split(
        tagged_ciphertexts, multiprocessing.cpu_count()
    )

    compound_pd = []
    compound_pd2 = []
    for teller in tellers:
        q1 = multiprocessing.Queue()
        q2 = multiprocessing.Queue()
        q3 = multiprocessing.Queue()
        processes = [
            multiprocessing.Process(
                target=teller.mp_partial_decrypt, args=(ciph, q1, q2, q3)
            )
            for ciph in split_ciphertexts
        ]
        for p in processes:
            p.daemon = True
            p.start()
        data = []
        data2 = []
        proofs = []
        for p in processes:
            data = data + q1.get()
            data2 = data2 + q2.get()
            proofs.append(q3.get())
        for p in processes:
            p.join()
            p.close()
        compound_pd.append(data)
        compound_pd2.append(data2)

    final_pd = []
    final_pd2 = []

    for i in range(len(compound_pd[0])):
        temp = []
        temp.append(i)
        temp2 = []
        temp2.append(i)
        subtemp = []
        subtemp2 = []
        for item in compound_pd:
            sub_item = multi_dim_index(item, i)
            subtemp.append(sub_item[1])
        for item in compound_pd2:
            sub_item = multi_dim_index(item, i)
            subtemp2.append(sub_item[1])

        temp.append(subtemp)
        temp2.append(subtemp2)

        final_pd.append(temp)
        final_pd2.append(temp2)

    global decrypted
    split_ciphertexts = tellers[0].ciphertext_list_split(
        final_pd, multiprocessing.cpu_count()
    )
    processes = [
        multiprocessing.Process(
            target=tellers[0].mp_full_decrypt,
            args=(ciph, tagged_ciphertexts, 1, q1),
        )
        for ciph in split_ciphertexts
    ]
    for p in processes:
        p.daemon = True
        p.start()
    data = []
    for p in processes:
        data = data + q1.get()

    for p in processes:
        p.join()
        p.close()
    vote_list = data
    split_ciphertexts = tellers[0].ciphertext_list_split(
        final_pd2, multiprocessing.cpu_count()
    )
    processes = [
        multiprocessing.Process(
            target=tellers[0].mp_full_decrypt,
            args=(ciph, tagged_ciphertexts, 2, q1),
        )
        for ciph in split_ciphertexts
    ]
    for p in processes:
        p.daemon = True
        p.start()
    data = []
    for p in processes:
        data = data + q1.get()

    for p in processes:
        p.join()
        p.close()
    comm_list = data

    comm = None
    for item in vote_list:
        index = item[0]
        for subitem in comm_list:
            if subitem[0] == index:
                comm = subitem[1]
                break
        verification_bb.append(
            {"v": item[1], "tracker": comm}
        )
    t_decryption = time.time() - t_decryption_start

def notification():
    """The tallying phase of the protocol.
    The 'r_i' term that corresponds to each voter is encrypted under
    their public key and sent privately to said voter.
    """
    for teller in tellers:
        for voter in voters:
            g_ri = teller.get_notification_entry(voter.id)
            voter.notify(g_ri)


def verification():
    """The verification phase of the protocol.

    The program aborts if verification fails for any voter at
    this stage.
    """
    for i in range(0, len(voters)):
        voter = voters[i]
        index = find_index_by_id(voter.id, final_bb)
        t_verification_single_start = time.time()
        tracker = voter.retrieve_tracker(mpz(final_bb[index]["comm"]))
        entry = find_entry_by_tracker(tracker, verification_bb)
        if entry["v"] == voter.g_vote:
            for tracker in g_trackers:
                if tracker["g_tracker"] == entry["tracker"]:
                    pass
        else:
            print("Error: Verification failed for voter" + str(voter.id))
            exit()
        t_verification = time.time() - t_verification_single_start
        global t_verification_single
        t_verification_single = t_verification_single + t_verification


def coercion_mitigation():
    """The coercion mitigation mechanism.

    A single voter selects another vote from the bulletin board and
    produces a fake dual key such that verification reveals a fake vote.
    """
    voter = voters[0]
    index = find_index_by_id(voter.id, final_bb)
    beta_term = mpz(final_bb[index]["comm"])
    # target = None
    # pick a random entry in the vbb
    for entry in verification_bb:
        if entry["v"] != voter.g_vote:
            # target = entry
            break
    key_inverse = invert(voter.secret_trapdoor_key, group.q)
    fake_dual_key = divm(beta_term, mpz(entry["tracker"]), group.p)
    fake_dual_key = powmod(fake_dual_key, key_inverse, group.p)
    ege = ElGamalEncryption(group)
    ciphertext = [fake_dual_key, beta_term]
    fake_tracker = ege.decrypt(voter.secret_trapdoor_key, ciphertext)


def print_verification_bb():
    """Prints the contents of the final bulletin board to console.
    """
    table = Texttable()
    table.add_row(["Vote", "Commitment"])

    for item in verification_bb:
        temp_entry = item
        for tracker in g_trackers:
            if tracker["g_tracker"] == temp_entry["tracker"]:
                temp_entry["tracker"] = tracker["tracker"]
                break
        for vote in vote_map:
            if vote["g_vote"] == temp_entry["v"]:
                temp_entry["v"] = vote["vote"]
                break
        table.add_row([str(temp_entry["v"]), str(temp_entry["tracker"])])
    print(table.draw())
    print()


print("Selene: Voting with Transparent Verifiability and Coercion-Mitigation")
print()

print("Running trials...")



tracker_voter_pairs = []
poc_setup()

t_setup_start = time.time()
setup()
t_setup = str(time.time() - t_setup_start)

voting()

t_tallying_start = time.time()
tallying()
t_tallying = str(time.time() - t_tallying_start)

t_notification_start = time.time()
notification()
t_notification = str(time.time() - t_notification_start)

verification()

t_coercion_mitigation_start = time.time()
coercion_mitigation()
t_coercion_mitigation = str(time.time() - t_coercion_mitigation_start)

print_verification_bb()

t_voting_single = str(t_voting_single / num_voters)
t_verification_single = str(t_verification_single / num_voters)


print()
print("Voter count: " + str(num_voters))
print("Tally teller count: " + str(num_tellers))

table = Texttable()
output_headings = [
    "Setup",
    "Voting (avg.)",
    "Tallying (Mixing)",
    "Tallying (Decryption)",
    "Notification",
    "Verification (avg.)",
    "Coercion Mitigation",
]

table.add_row(output_headings)
table.add_row(
    [
        t_setup,
        t_voting_single,
        t_mixing,
        t_decryption,
        t_notification,
        t_verification_single,
        t_coercion_mitigation,
    ]
)


print(table.draw())

file_name = "Selene-Timing-Data.xlsx"
if not Path(file_name).exists():
    results_workbook = Workbook()
    results_counter = 1
    results_max_row = 0
else:
    results_workbook = load_workbook(file_name)
    results_max_row = results_workbook.active.max_row
    results_counter = (
        results_workbook.active["A" + str(results_max_row)].value + 1
    )
results_worksheet = results_workbook.active


if results_max_row == 0:
    results_worksheet.append(
        [
            "N",
            "Voters",
            "Tellers",
            "Threshold",
            "Setup",
            "Voting (avg.)",
            "Tallying (Mixing)",
            "Tallying (Decryption)",
            "Notification",
            "Verification (avg.)",
            "Coercion Mitigation",
        ]
    )
timing_data = [
    results_counter,
    int(num_voters),
    int(num_tellers),
    int(k),
    float(t_setup),
    float(t_voting_single),
    float(t_mixing),
    float(t_decryption),
    float(t_notification),
    float(t_verification_single),
    float(t_coercion_mitigation),
]
results_worksheet.append(timing_data)
results_workbook.save(file_name)

print(
    '\nThese values (all in seconds) have been written to a file named "Selene-Timing-Data.xlsx", in the current working directory.'
)
